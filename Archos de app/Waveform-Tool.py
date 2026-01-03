# -*- coding: utf-8 -*-
"""
Waveform Manager / Designer (OTO)
- Editor basado en gráfica (sin tabla pesada)
- Herramientas: Invertir, Shift, Copiar/Pegar tramo, Rellenar, Patrón (#1/#0 -> generar -> aplicar)
- UNDO (Ctrl+Z) por snapshots
- Imagen de referencia: calibración eje X + medición de diente (2 clics)
- Librería: escaneo de carpetas, base SQLite, filtros, previsualización de waveform al seleccionar,
  cargar al editor, editar metadata, borrar múltiple
- Configuración: tema claro / oscuro

Requisitos:
    pip install matplotlib pillow openpyxl
"""

import os
import re
import csv
import json
import time
import sqlite3
import traceback
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ----- Dependencias opcionales con mensajes claros -----
try:
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
    from matplotlib.figure import Figure
except Exception:
    raise SystemExit(
        "Falta matplotlib.\n\nInstala con:\n    pip install matplotlib\n"
    )

try:
    from PIL import Image, ImageTk
except Exception:
    raise SystemExit(
        "Falta Pillow.\n\nInstala con:\n    pip install pillow\n"
    )

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None  # Export Excel quedará deshabilitado


TOTAL_DEGREES = 720

CHANNELS = [
    "A34",
    "A16",
    "A35",
    "A17",
    "B34",
    "B16",
    "B35",
    "B17",
]


# ---------------------------- Utilidades OTO ----------------------------

def sanitize_bit(v) -> int:
    return 1 if str(v).strip() == "1" else 0


def read_oto_wavecolumns(path: str) -> Dict[int, str]:
    """
    Regresa dict: {col_idx: bitstring}, donde col_idx inicia en 1.
    """
    out: Dict[int, str] = {}
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if not line.startswith("waveColumn"):
                continue
            parts = line.strip().split(";", 1)
            if len(parts) != 2:
                continue
            name, data_str = parts
            try:
                idx = int(name.replace("waveColumn", ""))
            except ValueError:
                continue
            bitstr = "".join(c for c in data_str if c in "01")
            if bitstr:
                out[idx] = bitstr
    return out


def detect_base_len_from_oto(path: str) -> Optional[int]:
    try:
        cols = read_oto_wavecolumns(path)
        if not cols:
            return None
        first = next(iter(cols.values()))
        return len(first)
    except Exception:
        return None


def decode_to_720(bitstr: str) -> List[int]:
    """
    Convierte bitstring de longitud L a 720 valores.
    Soporta:
        - L == 720  -> 1 bit por grado
        - L == 1440 -> 2 bits por grado (si cualquiera es 1 => 1)
        - otro      -> re-muestreo simple
    """
    L = len(bitstr)
    if L == TOTAL_DEGREES:
        return [1 if bitstr[i] == "1" else 0 for i in range(TOTAL_DEGREES)]
    if L == 2 * TOTAL_DEGREES:
        out = []
        for i in range(TOTAL_DEGREES):
            chunk = bitstr[2 * i:2 * i + 2]
            out.append(1 if "1" in chunk else 0)
        return out

    out = []
    ratio = L / TOTAL_DEGREES
    for i in range(TOTAL_DEGREES):
        src = int(i * ratio)
        if src < 0:
            src = 0
        if src >= L:
            src = L - 1
        out.append(1 if bitstr[src] == "1" else 0)
    return out


def read_waveform_from_oto(path: str) -> Dict[str, List[int]]:
    cols = read_oto_wavecolumns(path)
    data = {ch: [0] * TOTAL_DEGREES for ch in CHANNELS}
    for idx, bitstr in cols.items():
        if 1 <= idx <= len(CHANNELS):
            ch = CHANNELS[idx - 1]
            data[ch] = decode_to_720(bitstr)
    return data


def write_waveform_to_oto(base_path: str, out_path: str, data_720: Dict[str, List[int]],
                          channel_enabled: Dict[str, bool]) -> None:
    """
    Escribe waveColumnX en out_path tomando como base el archivo base_path.
    - Respeta base_len de waveColumn existente (720/1440/otro).
    - waveCheckBox se actualiza según channel_enabled.
    """
    with open(base_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    base_len = None
    for line in lines:
        if line.startswith("waveColumn"):
            parts = line.strip().split(";", 1)
            if len(parts) == 2:
                bits = "".join(c for c in parts[1] if c in "01")
                if bits:
                    base_len = len(bits)
                    break
    if base_len is None:
        base_len = 2 * TOTAL_DEGREES

    def to_bitstring(bits720: List[int]) -> str:
        s720 = "".join("1" if v else "0" for v in bits720)
        if base_len == TOTAL_DEGREES:
            return s720
        if base_len == 2 * TOTAL_DEGREES:
            return "".join(b * 2 for b in s720)
        out = []
        ratio = base_len / TOTAL_DEGREES
        for k in range(base_len):
            src = int(k / ratio)
            if src < 0:
                src = 0
            if src >= TOTAL_DEGREES:
                src = TOTAL_DEGREES - 1
            out.append(s720[src])
        return "".join(out)

    for col_idx in range(1, len(CHANNELS) + 1):
        ch = CHANNELS[col_idx - 1]
        if ch not in data_720:
            continue
        if not channel_enabled.get(ch, True):
            continue
        out_bits = to_bitstring(data_720[ch])
        new_line = f"waveColumn{col_idx};{out_bits}\n"
        replaced = False
        for i, line in enumerate(lines):
            if line.startswith(f"waveColumn{col_idx}"):
                lines[i] = new_line
                replaced = True
                break
        if not replaced:
            lines.append(new_line)

    checkbox_idx = None
    for i, line in enumerate(lines):
        if line.startswith("waveCheckBox"):
            checkbox_idx = i
            break
    flags = ["True" if channel_enabled.get(ch, True) else "False" for ch in CHANNELS]
    checkbox_line = "waveCheckBox;" + ";".join(flags) + "\n"
    if checkbox_idx is not None:
        lines[checkbox_idx] = checkbox_line
    else:
        lines.insert(0, checkbox_line)

    names_idx = None
    for i, line in enumerate(lines):
        if line.startswith("waveNames"):
            names_idx = i
            break
    if names_idx is None:
        lines.insert(0, "waveNames;" + ";".join(CHANNELS) + "\n")

    with open(out_path, "w", encoding="utf-8") as f:
        f.writelines(lines)


# ---------------------------- Parser de metadata desde .OTO (NUEVO) ----------------------------

_YEAR_RE = re.compile(r"(\d{4}(?:\s*-\s*\d{4})?)")
KNOWN_MODULES = {"PCM", "ECM", "TCM", "ABS", "TIPM", "ECU", "NGC", "BCM", "IPM", "CCM"}

def _safe_str(x) -> str:
    return "" if x is None else str(x).strip()

def parse_oto_metadata_fast(oto_path: str) -> dict:
    """
    Lee rápido metadatos típicos dentro del .oto.
    Busca:
      InfoDesginInformationModuleInfo0;...;MANUF;YEAR;MODEL;ENGINE;MODULE;...
    """
    meta = {
        "profile": os.path.splitext(os.path.basename(oto_path))[0],
        "fabricante": "",
        "marca": "",
        "modelo": "",
        "year": "",
        "engine": "",
        "module": "",
    }

    try:
        with open(oto_path, "r", encoding="utf-8", errors="ignore") as f:
            for _ in range(1100):
                line = f.readline()
                if not line:
                    break
                line = line.strip()
                if line.startswith("InfoDesginInformationModuleInfo0;"):
                    cols = line.split(";")
                    # Por tu ejemplo: 3=Manuf, 4=Year, 5=Model, 6=Engine, 7=Module
                    if len(cols) > 7:
                        meta["fabricante"] = _safe_str(cols[3])
                        meta["year"] = _safe_str(cols[4])
                        meta["modelo"] = _safe_str(cols[5])
                        meta["engine"] = _safe_str(cols[6])
                        meta["module"] = _safe_str(cols[7])
                    break
    except Exception:
        return meta

    y = meta["year"]
    m = _YEAR_RE.search(y)
    if m:
        meta["year"] = m.group(1).replace(" ", "")
    else:
        meta["year"] = y.strip()

    # Normaliza module (si trae extra)
    mod = meta["module"].strip()
    if mod and len(mod) > 8:
        # intenta encontrar token conocido
        tokens = re.split(r"[\s,_/]+", mod.upper())
        for t in tokens:
            if t in KNOWN_MODULES:
                meta["module"] = t
                break

    return meta

def infer_meta_from_path(root: str, file_path: str) -> dict:
    """
    Heurística por ruta:
        root/<fabricante>/<marca>/<modelo>/<year>/<engine>/<module>/<file.oto>
    Si no existe esa estructura, intenta rescatar algo sin romper.
    """
    out = {"fabricante": "", "marca": "", "modelo": "", "year": "", "engine": "", "module": ""}
    try:
        rel = os.path.relpath(file_path, root)
    except Exception:
        rel = file_path
    parts = [p for p in rel.split(os.sep) if p and p not in (".", "..")]
    if len(parts) < 2:
        return out

    # quita file
    parts_wo_file = parts[:-1]
    if not parts_wo_file:
        return out

    # intento directo por posiciones
    if len(parts_wo_file) >= 1:
        out["fabricante"] = parts_wo_file[0]
    if len(parts_wo_file) >= 2:
        out["marca"] = parts_wo_file[1]
    if len(parts_wo_file) >= 3:
        out["modelo"] = parts_wo_file[2]

    # busca year en cualquier segmento
    for p in parts_wo_file:
        m = _YEAR_RE.search(p)
        if m:
            out["year"] = m.group(1).replace(" ", "")
            break

    # engine suele traer "L" o "Diesel"
    for p in parts_wo_file[::-1]:
        if re.search(r"(\d\.\d\s*L|\d\.\dL|DIESEL)", p, flags=re.I):
            out["engine"] = p
            break

    # module por tokens conocidos
    for p in parts_wo_file[::-1]:
        up = p.upper()
        if up in KNOWN_MODULES:
            out["module"] = up
            break

    return out

def merge_meta(file_meta: dict, path_meta: dict) -> dict:
    """
    file_meta tiene prioridad. path_meta rellena vacíos.
    """
    out = dict(file_meta)
    for k in ("fabricante", "marca", "modelo", "year", "engine", "module"):
        if not out.get(k):
            out[k] = _safe_str(path_meta.get(k))
    # marca default = fabricante si sigue vacío
    if not out.get("marca") and out.get("fabricante"):
        out["marca"] = out["fabricante"]
    return out


# ---------------------------- DB Librería (SQLite) ----------------------------

@dataclass
class WaveRecord:
    id: int
    profile: str
    fabricante: str
    marca: str
    modelo: str
    year: str
    engine: str
    module: str
    file_path: str
    file_name: str
    updated_ts: float


class WaveDB:
    def __init__(self, db_path: str):
        self.db_path = db_path
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        self._init()

    def _ensure_columns(self, cols_needed: Dict[str, str]):
        cur = self.conn.cursor()
        cur.execute("PRAGMA table_info(waveforms)")
        existing = {r[1] for r in cur.fetchall()}  # name
        for col, coltype in cols_needed.items():
            if col not in existing:
                cur.execute(f"ALTER TABLE waveforms ADD COLUMN {col} {coltype}")
        self.conn.commit()

    def _init(self):
        cur = self.conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS waveforms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile TEXT,
            fabricante TEXT,
            marca TEXT,
            modelo TEXT,
            year TEXT,
            engine TEXT,
            module TEXT,
            file_path TEXT UNIQUE,
            file_name TEXT,
            updated_ts REAL
        )
        """)
        self.conn.commit()

        # Migración suave por si ya existía DB vieja
        self._ensure_columns({
            "profile": "TEXT",
            "module": "TEXT",
        })

        cur.execute("CREATE INDEX IF NOT EXISTS idx_filters ON waveforms(fabricante, marca, modelo, year, engine, module)")
        self.conn.commit()

    def upsert(self, profile: str, fabricante: str, marca: str, modelo: str, year: str, engine: str, module: str, file_path: str):
        file_name = os.path.basename(file_path)
        ts = time.time()
        cur = self.conn.cursor()
        cur.execute("""
        INSERT INTO waveforms(profile, fabricante, marca, modelo, year, engine, module, file_path, file_name, updated_ts)
        VALUES(?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(file_path) DO UPDATE SET
            profile=excluded.profile,
            fabricante=excluded.fabricante,
            marca=excluded.marca,
            modelo=excluded.modelo,
            year=excluded.year,
            engine=excluded.engine,
            module=excluded.module,
            file_name=excluded.file_name,
            updated_ts=excluded.updated_ts
        """, (profile, fabricante, marca, modelo, year, engine, module, file_path, file_name, ts))
        self.conn.commit()

    def delete_ids(self, ids: List[int]):
        if not ids:
            return
        cur = self.conn.cursor()
        cur.execute(f"DELETE FROM waveforms WHERE id IN ({','.join('?' for _ in ids)})", ids)
        self.conn.commit()

    def update_record(self, rec_id: int, profile: str, fabricante: str, marca: str, modelo: str, year: str, engine: str, module: str):
        cur = self.conn.cursor()
        cur.execute("""
        UPDATE waveforms
        SET profile=?, fabricante=?, marca=?, modelo=?, year=?, engine=?, module=?, updated_ts=?
        WHERE id=?
        """, (profile, fabricante, marca, modelo, year, engine, module, time.time(), rec_id))
        self.conn.commit()

    def list_distinct(self, field: str, where: Dict[str, str]) -> List[str]:
        allowed = {"fabricante", "marca", "modelo", "year", "engine", "module"}
        if field not in allowed:
            return []
        clauses = []
        params = []
        for k, v in where.items():
            if k in allowed and v:
                clauses.append(f"{k}=?")
                params.append(v)
        wsql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur = self.conn.cursor()
        cur.execute(f"SELECT DISTINCT {field} AS v FROM waveforms {wsql} ORDER BY v COLLATE NOCASE", params)
        return [r["v"] for r in cur.fetchall() if r["v"] is not None]

    def query(self, fabricante: str = "", marca: str = "", modelo: str = "", year: str = "", engine: str = "", module: str = "",
              text: str = "") -> List[WaveRecord]:
        clauses = []
        params = []
        if fabricante:
            clauses.append("fabricante=?")
            params.append(fabricante)
        if marca:
            clauses.append("marca=?")
            params.append(marca)
        if modelo:
            clauses.append("modelo=?")
            params.append(modelo)
        if year:
            clauses.append("year=?")
            params.append(year)
        if engine:
            clauses.append("engine=?")
            params.append(engine)
        if module:
            clauses.append("module=?")
            params.append(module)

        if text.strip():
            t = f"%{text.strip()}%"
            clauses.append("(file_name LIKE ? OR file_path LIKE ? OR modelo LIKE ? OR engine LIKE ? OR fabricante LIKE ? OR marca LIKE ? OR profile LIKE ? OR module LIKE ?)")
            params.extend([t, t, t, t, t, t, t, t])

        wsql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        cur = self.conn.cursor()
        cur.execute(
            f"SELECT * FROM waveforms {wsql} ORDER BY fabricante, marca, modelo, year, engine, module, profile, file_name",
            params
        )
        out = []
        for r in cur.fetchall():
            out.append(WaveRecord(
                id=r["id"],
                profile=r["profile"] or (r["file_name"] or ""),
                fabricante=r["fabricante"] or "",
                marca=r["marca"] or "",
                modelo=r["modelo"] or "",
                year=r["year"] or "",
                engine=r["engine"] or "",
                module=r["module"] or "",
                file_path=r["file_path"],
                file_name=r["file_name"] or os.path.basename(r["file_path"]),
                updated_ts=float(r["updated_ts"] or 0.0),
            ))
        return out

    def close(self):
        try:
            self.conn.close()
        except Exception:
            pass


# ---------------------------- Tema (claro/oscuro) ----------------------------

class ThemeManager:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.style = ttk.Style(self.root)
        self.theme_var = tk.StringVar(value="Oscuro")  # "Oscuro" / "Claro"

        self._dark = {
            "bg": "#1f1f1f",
            "fg": "#e8e8e8",
            "panel": "#2a2a2a",
            "entry_bg": "#2d2d2d",
            "entry_fg": "#ffffff",
            "accent": "#3d85c6",
            "tree_sel": "#2f5f9e",
        }
        self._light = {
            "bg": "#f3f3f3",
            "fg": "#111111",
            "panel": "#ffffff",
            "entry_bg": "#ffffff",
            "entry_fg": "#111111",
            "accent": "#1f77b4",
            "tree_sel": "#cfe8ff",
        }

    def colors(self) -> Dict[str, str]:
        return self._dark if self.theme_var.get() == "Oscuro" else self._light

    def apply(self):
        c = self.colors()

        try:
            self.style.theme_use("clam")
        except Exception:
            pass

        self.root.configure(bg=c["bg"])

        self.style.configure(".", background=c["bg"], foreground=c["fg"])
        self.style.configure("TFrame", background=c["bg"])
        self.style.configure("TLabel", background=c["bg"], foreground=c["fg"])
        self.style.configure("TLabelframe", background=c["bg"], foreground=c["fg"])
        self.style.configure("TLabelframe.Label", background=c["bg"], foreground=c["fg"])

        self.style.configure("TButton", padding=6)
        self.style.map("TButton",
                       background=[("active", c["panel"]), ("!active", c["panel"])],
                       foreground=[("active", c["fg"]), ("!active", c["fg"])])

        self.style.configure("TEntry",
                             fieldbackground=c["entry_bg"],
                             background=c["entry_bg"],
                             foreground=c["entry_fg"],
                             insertcolor=c["entry_fg"])

        self.style.configure("TCombobox",
                             fieldbackground=c["entry_bg"],
                             background=c["entry_bg"],
                             foreground=c["entry_fg"])

        self.style.configure("Treeview",
                             background=c["panel"],
                             fieldbackground=c["panel"],
                             foreground=c["fg"],
                             borderwidth=0)
        self.style.map("Treeview",
                       background=[("selected", c["tree_sel"])],
                       foreground=[("selected", "#ffffff")])

        self.style.configure("TNotebook", background=c["bg"], borderwidth=0)
        self.style.configure("TNotebook.Tab", padding=(12, 7))
        self.style.map("TNotebook.Tab",
                       background=[("selected", c["panel"]), ("!selected", c["bg"])],
                       foreground=[("selected", c["fg"]), ("!selected", c["fg"])])


# ---------------------------- Editor (solo gráfica + tools) ----------------------------

class WaveformEditorModel:
    def __init__(self):
        self.data: Dict[str, List[int]] = {ch: [0] * TOTAL_DEGREES for ch in CHANNELS}
        self.channel_enabled: Dict[str, bool] = {ch: True for ch in CHANNELS}
        self.selected_channel: str = CHANNELS[0]
        self.last_oto_path: Optional[str] = None

        self.undo_stack: List[Dict] = []
        self.undo_limit = 60
        self._in_undo = False

    def push_undo_snapshot(self):
        if self._in_undo:
            return
        snap = {ch: self.data[ch][:] for ch in CHANNELS}
        self.undo_stack.append({"type": "snapshot", "data": snap})
        if len(self.undo_stack) > self.undo_limit:
            self.undo_stack.pop(0)

    def undo(self) -> bool:
        if not self.undo_stack:
            return False
        self._in_undo = True
        try:
            act = self.undo_stack.pop()
            if act["type"] == "snapshot":
                d = act["data"]
                for ch in CHANNELS:
                    self.data[ch] = d.get(ch, [0] * TOTAL_DEGREES)[:]
                return True
        finally:
            self._in_undo = False
        return False

    def invert_channel(self, ch: str):
        self.push_undo_snapshot()
        self.data[ch] = [0 if v else 1 for v in self.data[ch]]

    def shift_channel(self, ch: str, amount: int, direction: int):
        if amount == 0:
            return
        self.push_undo_snapshot()
        n = TOTAL_DEGREES
        k = (amount * direction) % n
        vals = self.data[ch]
        self.data[ch] = vals[-k:] + vals[:-k]

    def copy_segment(self, ch: str, start: int, end: int) -> List[int]:
        if end < start:
            start, end = end, start
        start = max(0, min(TOTAL_DEGREES - 1, start))
        end = max(0, min(TOTAL_DEGREES - 1, end))
        return self.data[ch][start:end + 1]

    def paste_segment_circular(self, ch: str, start: int, seg: List[int]):
        self.push_undo_snapshot()
        start = max(0, min(TOTAL_DEGREES - 1, start))
        for i, v in enumerate(seg):
            idx = (start + i) % TOTAL_DEGREES
            self.data[ch][idx] = 1 if v else 0

    def fill_range(self, ch: str, start: int, end: int, val: int):
        self.push_undo_snapshot()
        if end < start:
            start, end = end, start
        start = max(0, min(TOTAL_DEGREES - 1, start))
        end = max(0, min(TOTAL_DEGREES - 1, end))
        v = 1 if val else 0
        for i in range(start, end + 1):
            self.data[ch][i] = v

    def apply_pattern(self, ch: str, start: int, end: int, pattern: str):
        self.push_undo_snapshot()
        if end < start:
            start, end = end, start
        start = max(0, min(TOTAL_DEGREES - 1, start))
        end = max(0, min(TOTAL_DEGREES - 1, end))
        p = [1 if c == "1" else 0 for c in pattern if c in ("0", "1")]
        if not p:
            return
        L = len(p)
        for k, deg in enumerate(range(start, end + 1)):
            self.data[ch][deg] = p[k % L]

    def toggle_point(self, ch: str, deg: int):
        if deg < 0 or deg >= TOTAL_DEGREES:
            return
        self.push_undo_snapshot()
        self.data[ch][deg] = 0 if self.data[ch][deg] else 1


class PlotAndToolsTab(ttk.Frame):
    def __init__(self, master, app, theme: ThemeManager):
        super().__init__(master)
        self.app = app
        self.theme = theme

        self.model = WaveformEditorModel()
        self.segment_clip: Optional[List[int]] = None

        self._build_ui()
        self._bind_shortcuts()
        self.redraw_plot()

    def _build_ui(self):
        outer = ttk.Frame(self)
        outer.pack(fill="both", expand=True, padx=10, pady=10)

        pan = ttk.Panedwindow(outer, orient="horizontal")
        pan.pack(fill="both", expand=True)

        tools = ttk.Frame(pan)
        pan.add(tools, weight=0)

        plotp = ttk.Frame(pan)
        pan.add(plotp, weight=1)

        lf1 = ttk.Labelframe(tools, text="Canales")
        lf1.pack(fill="x", pady=(0, 10))

        row = ttk.Frame(lf1)
        row.pack(fill="x", padx=8, pady=6)

        ttk.Label(row, text="Canal activo:").pack(side="left")
        self.sel_ch_var = tk.StringVar(value=self.model.selected_channel)
        self.sel_ch_cb = ttk.Combobox(row, textvariable=self.sel_ch_var, state="readonly",
                                      values=CHANNELS, width=8)
        self.sel_ch_cb.pack(side="left", padx=6)
        self.sel_ch_cb.bind("<<ComboboxSelected>>", lambda e: self._on_select_channel())

        grid = ttk.Frame(lf1)
        grid.pack(fill="x", padx=8, pady=(0, 8))
        self.en_vars: Dict[str, tk.BooleanVar] = {}
        for i, ch in enumerate(CHANNELS):
            v = tk.BooleanVar(value=True)
            self.en_vars[ch] = v
            cb = ttk.Checkbutton(grid, text=ch, variable=v, command=self._on_toggle_channels)
            cb.grid(row=i // 4, column=i % 4, sticky="w", padx=6, pady=2)

        lf2 = ttk.Labelframe(tools, text="Acciones")
        lf2.pack(fill="x", pady=(0, 10))

        r = ttk.Frame(lf2)
        r.pack(fill="x", padx=8, pady=8)

        ttk.Button(r, text="Invertir (0↔1)", command=self._invert).pack(fill="x", pady=2)
        ttk.Button(r, text="Undo (Ctrl+Z)", command=self._undo).pack(fill="x", pady=2)

        sh = ttk.Frame(lf2)
        sh.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(sh, text="Shift grados:").pack(side="left")
        self.shift_var = tk.StringVar(value="10")
        ttk.Entry(sh, textvariable=self.shift_var, width=6).pack(side="left", padx=6)
        ttk.Button(sh, text="Shift +", command=lambda: self._shift(1)).pack(side="left", padx=3)
        ttk.Button(sh, text="Shift -", command=lambda: self._shift(-1)).pack(side="left", padx=3)

        lf3 = ttk.Labelframe(tools, text="Segmento (Copiar / Pegar)")
        lf3.pack(fill="x", pady=(0, 10))
        seg = ttk.Frame(lf3)
        seg.pack(fill="x", padx=8, pady=8)

        self.copy_from = tk.StringVar(value="0")
        self.copy_to = tk.StringVar(value="20")
        self.paste_at = tk.StringVar(value="0")

        a = ttk.Frame(seg)
        a.pack(fill="x", pady=2)
        ttk.Label(a, text="Desde:").pack(side="left")
        ttk.Entry(a, textvariable=self.copy_from, width=6).pack(side="left", padx=6)
        ttk.Label(a, text="Hasta:").pack(side="left")
        ttk.Entry(a, textvariable=self.copy_to, width=6).pack(side="left", padx=6)
        ttk.Button(a, text="Copiar", command=self._copy_seg).pack(side="left", padx=6)

        b = ttk.Frame(seg)
        b.pack(fill="x", pady=2)
        ttk.Label(b, text="Pegar en:").pack(side="left")
        ttk.Entry(b, textvariable=self.paste_at, width=6).pack(side="left", padx=6)
        ttk.Button(b, text="Pegar (circular)", command=self._paste_seg).pack(side="left", padx=6)

        lf4 = ttk.Labelframe(tools, text="Rellenar rango")
        lf4.pack(fill="x", pady=(0, 10))
        fr = ttk.Frame(lf4)
        fr.pack(fill="x", padx=8, pady=8)

        self.fill_val = tk.StringVar(value="1")
        self.fill_from = tk.StringVar(value="0")
        self.fill_to = tk.StringVar(value="20")

        r1 = ttk.Frame(fr)
        r1.pack(fill="x", pady=2)
        ttk.Label(r1, text="Valor (0/1):").pack(side="left")
        ttk.Entry(r1, textvariable=self.fill_val, width=4).pack(side="left", padx=6)

        r2 = ttk.Frame(fr)
        r2.pack(fill="x", pady=2)
        ttk.Label(r2, text="Desde:").pack(side="left")
        ttk.Entry(r2, textvariable=self.fill_from, width=6).pack(side="left", padx=6)
        ttk.Label(r2, text="Hasta:").pack(side="left")
        ttk.Entry(r2, textvariable=self.fill_to, width=6).pack(side="left", padx=6)
        ttk.Button(r2, text="Aplicar", command=self._fill).pack(side="left", padx=6)

        lf5 = ttk.Labelframe(tools, text="Patrón")
        lf5.pack(fill="x", pady=(0, 10))
        pr = ttk.Frame(lf5)
        pr.pack(fill="x", padx=8, pady=8)

        self.pat_ones = tk.StringVar(value="4")
        self.pat_zeros = tk.StringVar(value="2")
        self.pat_str = tk.StringVar(value="111100")
        self.pat_from = tk.StringVar(value="0")
        self.pat_to = tk.StringVar(value="719")

        p1 = ttk.Frame(pr)
        p1.pack(fill="x", pady=2)
        ttk.Label(p1, text="#1:").pack(side="left")
        ttk.Entry(p1, textvariable=self.pat_ones, width=6).pack(side="left", padx=6)
        ttk.Label(p1, text="#0:").pack(side="left")
        ttk.Entry(p1, textvariable=self.pat_zeros, width=6).pack(side="left", padx=6)
        ttk.Button(p1, text="Generar", command=self._gen_pattern).pack(side="left", padx=6)

        p2 = ttk.Frame(pr)
        p2.pack(fill="x", pady=2)
        ttk.Label(p2, text="Patrón (0/1):").pack(side="left")
        ttk.Entry(p2, textvariable=self.pat_str, width=12).pack(side="left", padx=6)

        p3 = ttk.Frame(pr)
        p3.pack(fill="x", pady=2)
        ttk.Label(p3, text="Desde:").pack(side="left")
        ttk.Entry(p3, textvariable=self.pat_from, width=6).pack(side="left", padx=6)
        ttk.Label(p3, text="Hasta:").pack(side="left")
        ttk.Entry(p3, textvariable=self.pat_to, width=6).pack(side="left", padx=6)
        ttk.Button(p3, text="Aplicar patrón", command=self._apply_pattern).pack(side="left", padx=6)

        lf6 = ttk.Labelframe(tools, text="Archivos")
        lf6.pack(fill="x")

        af = ttk.Frame(lf6)
        af.pack(fill="x", padx=8, pady=8)
        ttk.Button(af, text="Importar .oto (Editor)", command=self.on_import_oto).pack(fill="x", pady=2)
        ttk.Button(af, text="Escribir waveform en .oto", command=self.on_write_oto).pack(fill="x", pady=2)
        ttk.Button(af, text="Exportar Excel NG (.xlsx)", command=self.on_export_excel).pack(fill="x", pady=2)

        lfplot = ttk.Labelframe(plotp, text="Gráfica + Tools")
        lfplot.pack(fill="both", expand=True)

        self.fig = Figure(figsize=(7.5, 4.8), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=lfplot)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

        tb = ttk.Frame(lfplot)
        tb.pack(fill="x")
        self.toolbar = NavigationToolbar2Tk(self.canvas, tb)
        self.toolbar.update()

        self.status_var = tk.StringVar(value="Grado: -")
        ttk.Label(lfplot, textvariable=self.status_var).pack(anchor="w", padx=8, pady=(0, 8))

        self.cursor_line = None
        self.canvas.mpl_connect("motion_notify_event", self._on_plot_hover)
        self.canvas.mpl_connect("button_press_event", self._on_plot_click)

    def _bind_shortcuts(self):
        self.app.bind_all("<Control-z>", lambda e: self._undo())
        self.app.bind_all("<Control-Z>", lambda e: self._undo())

    def _on_select_channel(self):
        ch = self.sel_ch_var.get().strip()
        if ch in CHANNELS:
            self.model.selected_channel = ch

    def _on_toggle_channels(self):
        enabled_any = any(v.get() for v in self.en_vars.values())
        if not enabled_any:
            messagebox.showwarning("Canales", "Debes dejar al menos 1 canal habilitado.")
            self.en_vars[CHANNELS[0]].set(True)

        for ch in CHANNELS:
            self.model.channel_enabled[ch] = bool(self.en_vars[ch].get())

        if not self.model.channel_enabled.get(self.model.selected_channel, True):
            for ch in CHANNELS:
                if self.model.channel_enabled[ch]:
                    self.model.selected_channel = ch
                    self.sel_ch_var.set(ch)
                    break

        self.redraw_plot()

    def _require_channel(self) -> Optional[str]:
        ch = self.model.selected_channel
        if not ch or ch not in CHANNELS:
            messagebox.showerror("Error", "No hay canal válido seleccionado.")
            return None
        if not self.model.channel_enabled.get(ch, True):
            messagebox.showerror("Error", f"El canal '{ch}' está deshabilitado.")
            return None
        return ch

    def _invert(self):
        ch = self._require_channel()
        if not ch:
            return
        self.model.invert_channel(ch)
        self.redraw_plot()

    def _shift(self, direction: int):
        ch = self._require_channel()
        if not ch:
            return
        try:
            amount = int(self.shift_var.get())
        except ValueError:
            messagebox.showerror("Error", "Shift inválido.")
            return
        self.model.shift_channel(ch, amount, direction)
        self.redraw_plot()

    def _undo(self):
        if self.model.undo():
            self.redraw_plot()

    def _copy_seg(self):
        ch = self._require_channel()
        if not ch:
            return
        try:
            a = int(self.copy_from.get())
            b = int(self.copy_to.get())
        except ValueError:
            messagebox.showerror("Error", "Desde/Hasta inválidos.")
            return
        self.segment_clip = self.model.copy_segment(ch, a, b)
        messagebox.showinfo("Copiar", f"Tramo copiado ({len(self.segment_clip)} pts) en canal {ch}.")

    def _paste_seg(self):
        ch = self._require_channel()
        if not ch:
            return
        if not self.segment_clip:
            messagebox.showerror("Error", "No hay tramo copiado.")
            return
        try:
            start = int(self.paste_at.get())
        except ValueError:
            messagebox.showerror("Error", "Inicio inválido.")
            return
        self.model.paste_segment_circular(ch, start, self.segment_clip)
        self.redraw_plot()

    def _fill(self):
        ch = self._require_channel()
        if not ch:
            return
        try:
            v = 1 if str(self.fill_val.get()).strip() == "1" else 0
            a = int(self.fill_from.get())
            b = int(self.fill_to.get())
        except ValueError:
            messagebox.showerror("Error", "Rango inválido.")
            return
        self.model.fill_range(ch, a, b, v)
        self.redraw_plot()

    def _gen_pattern(self):
        try:
            n1 = int(self.pat_ones.get())
            n0 = int(self.pat_zeros.get())
        except ValueError:
            messagebox.showerror("Error", "#1 y #0 deben ser enteros.")
            return
        if n1 < 0 or n0 < 0 or (n1 == 0 and n0 == 0):
            messagebox.showerror("Error", "Valores inválidos para patrón.")
            return
        self.pat_str.set(("1" * n1) + ("0" * n0))

    def _apply_pattern(self):
        ch = self._require_channel()
        if not ch:
            return
        pat = self.pat_str.get().strip()
        if not pat or any(c not in "01" for c in pat):
            messagebox.showerror("Error", "Patrón inválido. Solo 0 y 1.")
            return
        try:
            a = int(self.pat_from.get())
            b = int(self.pat_to.get())
        except ValueError:
            messagebox.showerror("Error", "Desde/Hasta inválidos.")
            return
        self.model.apply_pattern(ch, a, b, pat)
        self.redraw_plot()

    def _on_plot_hover(self, event):
        if event is None or event.inaxes != self.ax or event.xdata is None:
            self.status_var.set("Grado: -")
            if self.cursor_line is not None:
                self.cursor_line.set_visible(False)
                self.canvas.draw_idle()
            return

        deg = int(round(event.xdata))
        deg = max(0, min(TOTAL_DEGREES - 1, deg))

        parts = [f"Grado: {deg}"]
        for ch in CHANNELS:
            if self.model.channel_enabled.get(ch, True):
                parts.append(f"{ch}={self.model.data[ch][deg]}")
        self.status_var.set(" | ".join(parts))

        if self.cursor_line is None:
            self.cursor_line = self.ax.axvline(deg, linewidth=1.0)
        else:
            self.cursor_line.set_xdata([deg, deg])
            self.cursor_line.set_visible(True)

        self.canvas.draw_idle()

    def _on_plot_click(self, event):
        if event is None or event.inaxes != self.ax or event.xdata is None:
            return
        if event.button != 1:
            return
        ch = self._require_channel()
        if not ch:
            return
        deg = int(round(event.xdata))
        deg = max(0, min(TOTAL_DEGREES - 1, deg))
        self.model.toggle_point(ch, deg)
        self.redraw_plot()

    def redraw_plot(self):
        c = self.theme.colors()

        self.ax.clear()
        self.fig.patch.set_facecolor(c["panel"])
        self.ax.set_facecolor(c["panel"])

        degs = list(range(TOTAL_DEGREES))
        offset = 0.0
        offset_step = 1.2
        plotted = 0

        for ch in CHANNELS:
            if not self.model.channel_enabled.get(ch, True):
                continue
            vals = [v + offset for v in self.model.data[ch]]
            self.ax.plot(degs, vals, label=ch)
            plotted += 1
            offset += offset_step

        self.ax.set_xlabel("Degree")
        self.ax.set_ylabel("Nivel (desplazado)")
        self.ax.set_xlim(0, TOTAL_DEGREES - 1)

        xt = list(range(0, TOTAL_DEGREES, 50))
        if (TOTAL_DEGREES - 1) not in xt:
            xt.append(TOTAL_DEGREES - 1)
        self.ax.set_xticks(xt)

        if plotted > 0:
            max_offset = (plotted - 1) * offset_step
            self.ax.set_ylim(-0.5, max_offset + 1.8)
            self.ax.legend(loc="upper right", fontsize="x-small")

        self.cursor_line = None
        self.canvas.draw_idle()

    def on_import_oto(self):
        path = filedialog.askopenfilename(
            title="Seleccionar .oto",
            filetypes=[("OTO", "*.oto"), ("Todos", "*.*")]
        )
        if not path:
            return
        try:
            self.model.push_undo_snapshot()
            self.model.data = read_waveform_from_oto(path)
            self.model.last_oto_path = path
            self.redraw_plot()
            messagebox.showinfo("Importar", f"Importado:\n{os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Importar", f"No se pudo importar.\n\n{e}")

    def on_write_oto(self):
        base = filedialog.askopenfilename(
            title="Seleccionar .oto base",
            filetypes=[("OTO", "*.oto"), ("Todos", "*.*")]
        )
        if not base:
            return
        suggested = os.path.splitext(base)[0] + "_waveform.oto"
        out = filedialog.asksaveasfilename(
            title="Guardar .oto",
            defaultextension=".oto",
            initialfile=os.path.basename(suggested),
            filetypes=[("OTO", "*.oto"), ("Todos", "*.*")]
        )
        if not out:
            return
        try:
            write_waveform_to_oto(
                base_path=base,
                out_path=out,
                data_720=self.model.data,
                channel_enabled=self.model.channel_enabled
            )
            self.model.last_oto_path = out
            messagebox.showinfo("Guardar", f"Guardado:\n{out}")
        except Exception as e:
            messagebox.showerror("Guardar", f"No se pudo escribir.\n\n{e}")

    def on_export_excel(self):
        if Workbook is None:
            messagebox.showerror("Excel", "Falta openpyxl. Instala con:\n\npip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            title="Guardar Excel NG (.xlsx)",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if not path:
            return

        try:
            data = self.model.data

            net_to_channel = {
                "A34": "A34",
                "A16": "A16",
                "A35": "A35",
                "A17": "A17",
                "B34": "B34",
                "B16": "B16",
                "B35": "B35",
                "A15&B17": "B17",
            }
            header_nets = ["A34", "A16", "A35", "A17", "B34", "B16", "B35", "A15&B17"]
            header_alias = ["sync", "crank1", "cam0", "", "crank2", "cam2", "", ""]

            wb = Workbook()
            ws = wb.active
            ws.title = "waveform"

            ws.cell(row=7, column=3, value="sync")
            ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=5)

            ws.cell(row=8, column=2, value="degree")
            for i, net in enumerate(header_nets):
                ws.cell(row=8, column=3 + i, value=net)
            ws.cell(row=8, column=3 + len(header_nets), value="byte")

            for i, alias in enumerate(header_alias):
                ws.cell(row=9, column=3 + i, value=alias)
            ws.cell(row=9, column=3 + len(header_nets), value="byte")

            start_row = 10
            for deg in range(TOTAL_DEGREES):
                excel_row = start_row + deg
                ws.cell(row=excel_row, column=2, value=deg)
                bits_for_byte = []
                for i, net in enumerate(header_nets):
                    ch = net_to_channel.get(net)
                    bit = 1 if (ch and data[ch][deg]) else 0
                    ws.cell(row=excel_row, column=3 + i, value=bit)
                    bits_for_byte.append(bit)

                byte_val = 0
                for bit_index, bit in enumerate(bits_for_byte):
                    if bit:
                        byte_val |= (1 << bit_index)
                ws.cell(row=excel_row, column=3 + len(header_nets), value=byte_val)

            wb.save(path)
            messagebox.showinfo("Excel", f"Exportado:\n{path}")
        except Exception as e:
            messagebox.showerror("Excel", f"No se pudo exportar.\n\n{e}")


# ---------------------------- Tab Imagen (calibración y medición) ----------------------------

class ImageTab(ttk.Frame):
    def __init__(self, master, app, theme: ThemeManager):
        super().__init__(master)
        self.app = app
        self.theme = theme

        self.ref_image = None
        self.ref_image_tk = None

        self.digitize_mode = "idle"  # calibrate / measure_tooth / idle
        self.calibration_points: List[int] = []
        self.tooth_points: List[int] = []
        self.px_per_deg: Optional[float] = None
        self.cal_deg0: Optional[float] = None
        self.cal_x0: Optional[int] = None

        self._build_ui()

    def _build_ui(self):
        outer = ttk.Frame(self)
        outer.pack(fill="both", expand=True, padx=10, pady=10)

        top = ttk.Frame(outer)
        top.pack(fill="x")

        ttk.Button(top, text="Cargar imagen (referencia)", command=self.load_image).pack(side="left")

        self.cal_status_var = tk.StringVar(value="Sin calibrar")
        ttk.Label(top, textvariable=self.cal_status_var).pack(side="left", padx=12)

        lf = ttk.Labelframe(outer, text="Imagen")
        lf.pack(fill="both", expand=True, pady=(10, 0))

        self.image_label = ttk.Label(lf)
        self.image_label.pack(fill="both", expand=True)
        self.image_label.bind("<Button-1>", self.on_image_click)

        ctrl = ttk.Labelframe(outer, text="Herramientas")
        ctrl.pack(fill="x", pady=(10, 0))

        row1 = ttk.Frame(ctrl)
        row1.pack(fill="x", padx=8, pady=6)

        ttk.Label(row1, text="Cal X grado 1:").pack(side="left")
        self.cal_deg1_var = tk.StringVar(value="0")
        ttk.Entry(row1, textvariable=self.cal_deg1_var, width=8).pack(side="left", padx=6)

        ttk.Label(row1, text="grado 2:").pack(side="left")
        self.cal_deg2_var = tk.StringVar(value="720")
        ttk.Entry(row1, textvariable=self.cal_deg2_var, width=8).pack(side="left", padx=6)

        ttk.Button(row1, text="Iniciar calibración (2 clics)", command=self.start_calibration).pack(side="left", padx=8)

        row2 = ttk.Frame(ctrl)
        row2.pack(fill="x", padx=8, pady=(0, 8))

        ttk.Button(row2, text="Medir diente (2 clics)", command=self.start_tooth_measure).pack(side="left")

        self.tooth_var = tk.StringVar(value="Ancho diente: -")
        ttk.Label(row2, textvariable=self.tooth_var).pack(side="left", padx=12)

    def load_image(self):
        path = filedialog.askopenfilename(
            title="Seleccionar imagen",
            filetypes=[("Imágenes", "*.png *.jpg *.jpeg *.bmp *.gif"), ("Todos", "*.*")]
        )
        if not path:
            return
        try:
            img = Image.open(path)
        except Exception as e:
            messagebox.showerror("Imagen", f"No se pudo abrir.\n\n{e}")
            return

        max_w, max_h = 1400, 900
        img.thumbnail((max_w, max_h), Image.LANCZOS)

        self.ref_image = img
        self.ref_image_tk = ImageTk.PhotoImage(img)
        self.image_label.configure(image=self.ref_image_tk)

        self.digitize_mode = "idle"
        self.calibration_points = []
        self.tooth_points = []
        self.px_per_deg = None
        self.cal_deg0 = None
        self.cal_x0 = None
        self.cal_status_var.set("Sin calibrar")
        self.tooth_var.set("Ancho diente: -")

    def start_calibration(self):
        if not self.ref_image_tk:
            messagebox.showerror("Calibración", "Primero carga una imagen.")
            return
        try:
            d1 = float(self.cal_deg1_var.get())
            d2 = float(self.cal_deg2_var.get())
        except ValueError:
            messagebox.showerror("Calibración", "Grados inválidos.")
            return
        if d1 == d2:
            messagebox.showerror("Calibración", "grado 1 y grado 2 deben ser distintos.")
            return
        self.digitize_mode = "calibrate"
        self.calibration_points = []
        self.cal_status_var.set("Calibrando: clic 1/2 en eje X...")

    def start_tooth_measure(self):
        if self.px_per_deg in (None, 0) or self.cal_deg0 is None or self.cal_x0 is None:
            messagebox.showerror("Medición", "Primero calibra el eje X.")
            return
        self.digitize_mode = "measure_tooth"
        self.tooth_points = []
        self.tooth_var.set("Ancho diente: esperando clic 1/2...")

    def on_image_click(self, event):
        if not self.ref_image_tk:
            return
        x = int(event.x)

        if self.digitize_mode == "calibrate":
            self.calibration_points.append(x)
            if len(self.calibration_points) == 1:
                self.cal_status_var.set("Calibrando: clic 2/2 en eje X...")
                return
            if len(self.calibration_points) == 2:
                try:
                    d1 = float(self.cal_deg1_var.get())
                    d2 = float(self.cal_deg2_var.get())
                except ValueError:
                    messagebox.showerror("Calibración", "Grados inválidos.")
                    self.digitize_mode = "idle"
                    return
                x1, x2 = self.calibration_points
                if x1 == x2:
                    messagebox.showerror("Calibración", "Los 2 clics no pueden ser iguales.")
                    self.digitize_mode = "idle"
                    return
                self.cal_x0 = x1
                self.cal_deg0 = d1
                self.px_per_deg = (x2 - x1) / (d2 - d1)
                self.digitize_mode = "idle"
                self.calibration_points = []
                self.cal_status_var.set(f"Calibrado: {self.px_per_deg:.4f} px/deg")
                messagebox.showinfo("Calibración", f"OK\npx/deg = {self.px_per_deg:.6f}")
                return

        if self.digitize_mode == "measure_tooth":
            self.tooth_points.append(x)
            if len(self.tooth_points) == 1:
                self.tooth_var.set("Ancho diente: esperando clic 2/2...")
                return
            if len(self.tooth_points) == 2 and self.px_per_deg not in (None, 0):
                x1, x2 = self.tooth_points
                deg1 = self.cal_deg0 + (x1 - self.cal_x0) / self.px_per_deg
                deg2 = self.cal_deg0 + (x2 - self.cal_x0) / self.px_per_deg
                ancho = abs(deg2 - deg1)
                self.tooth_var.set(f"Ancho diente: {ancho:.2f}°")
                messagebox.showinfo("Medición", f"Ancho diente = {ancho:.4f}°")
                self.digitize_mode = "idle"
                self.tooth_points = []


# ---------------------------- Tab Librería (ACTUALIZADO estilo screenshot + preview) ----------------------------

class LibraryTab(ttk.Frame):
    def __init__(self, master, app, theme: ThemeManager, db: WaveDB, editor_tab: PlotAndToolsTab):
        super().__init__(master)
        self.app = app
        self.theme = theme
        self.db = db
        self.editor_tab = editor_tab

        self.root_dir = tk.StringVar(value=os.path.join(os.path.dirname(os.path.abspath(__file__)), "WaveformDB"))
        self.search_text = tk.StringVar(value="")

        self.fab_var = tk.StringVar(value="")
        self.marca_var = tk.StringVar(value="")
        self.modelo_var = tk.StringVar(value="")
        self.year_var = tk.StringVar(value="")
        self.engine_var = tk.StringVar(value="")
        self.module_var = tk.StringVar(value="")

        self.records: List[WaveRecord] = []
        self.selected_rec: Optional[WaveRecord] = None

        self._build_ui()
        self.refresh_filters()
        self.refresh_list()

    def _build_ui(self):
        outer = ttk.Frame(self)
        outer.pack(fill="both", expand=True, padx=10, pady=10)

        top = ttk.Labelframe(outer, text="Base / Escaneo")
        top.pack(fill="x")

        row = ttk.Frame(top)
        row.pack(fill="x", padx=8, pady=8)

        ttk.Label(row, text="Carpeta raíz:").pack(side="left")
        ttk.Entry(row, textvariable=self.root_dir).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(row, text="Seleccionar...", command=self.pick_root).pack(side="left", padx=4)
        ttk.Button(row, text="Escanear / Actualizar DB", command=self.scan_root).pack(side="left", padx=4)

        filt = ttk.Labelframe(outer, text="Filtros")
        filt.pack(fill="x", pady=(10, 0))

        fr = ttk.Frame(filt)
        fr.pack(fill="x", padx=8, pady=8)

        def combo(lbl, var, width=16):
            box = ttk.Frame(fr)
            ttk.Label(box, text=lbl).pack(anchor="w")
            cb = ttk.Combobox(box, textvariable=var, state="readonly", width=width)
            cb.pack(fill="x")
            return box, cb

        b1, self.cb_fab = combo("Manuf (Fabricante)", self.fab_var, 18)
        b2, self.cb_marca = combo("Make (Marca)", self.marca_var, 18)
        b3, self.cb_modelo = combo("Model (Modelo)", self.modelo_var, 20)
        b4, self.cb_year = combo("Year (Año)", self.year_var, 10)
        b5, self.cb_engine = combo("Engine", self.engine_var, 16)
        b6, self.cb_module = combo("Module", self.module_var, 12)

        b1.pack(side="left", padx=6)
        b2.pack(side="left", padx=6)
        b3.pack(side="left", padx=6)
        b4.pack(side="left", padx=6)
        b5.pack(side="left", padx=6)
        b6.pack(side="left", padx=6)

        for cb in [self.cb_fab, self.cb_marca, self.cb_modelo, self.cb_year, self.cb_engine, self.cb_module]:
            cb.bind("<<ComboboxSelected>>", lambda e: self.on_filter_change())

        sr = ttk.Frame(filt)
        sr.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(sr, text="Buscar texto:").pack(side="left")
        ttk.Entry(sr, textvariable=self.search_text).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(sr, text="Aplicar", command=self.refresh_list).pack(side="left")
        ttk.Button(sr, text="Limpiar", command=self.clear_filters).pack(side="left", padx=6)

        pan = ttk.Panedwindow(outer, orient="horizontal")
        pan.pack(fill="both", expand=True, pady=(10, 0))

        left = ttk.Frame(pan)
        right = ttk.Frame(pan)
        pan.add(left, weight=2)
        pan.add(right, weight=2)

        lft = ttk.Labelframe(left, text="Resultados (doble click = cargar al editor)")
        lft.pack(fill="both", expand=True)

        # Tabla estilo screenshot: Profile / Manuf / Year / Make / Model / Engine / Module / File
        cols = ("profile", "fabricante", "year", "marca", "modelo", "engine", "module", "file")
        self.tree = ttk.Treeview(lft, columns=cols, show="headings", selectmode="extended")

        self.tree.heading("profile", text="Profile")
        self.tree.heading("fabricante", text="Manuf")
        self.tree.heading("year", text="Year")
        self.tree.heading("marca", text="Make")
        self.tree.heading("modelo", text="Model")
        self.tree.heading("engine", text="Engine")
        self.tree.heading("module", text="Module")
        self.tree.heading("file", text="File")

        self.tree.column("profile", width=220, anchor="w")
        self.tree.column("fabricante", width=120, anchor="w")
        self.tree.column("year", width=80, anchor="center")
        self.tree.column("marca", width=130, anchor="w")
        self.tree.column("modelo", width=170, anchor="w")
        self.tree.column("engine", width=140, anchor="w")
        self.tree.column("module", width=90, anchor="w")
        self.tree.column("file", width=260, anchor="w")

        ysb = ttk.Scrollbar(lft, orient="vertical", command=self.tree.yview)
        xsb = ttk.Scrollbar(lft, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")

        lft.rowconfigure(0, weight=1)
        lft.columnconfigure(0, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self.on_select_row)
        self.tree.bind("<Double-1>", lambda e: self.load_selected_to_editor())

        btns = ttk.Frame(left)
        btns.pack(fill="x", pady=(8, 0))

        ttk.Button(btns, text="Cargar al Editor", command=self.load_selected_to_editor).pack(side="left")
        ttk.Button(btns, text="Editar metadata", command=self.edit_selected).pack(side="left", padx=6)
        ttk.Button(btns, text="Eliminar seleccionados", command=self.delete_selected).pack(side="left", padx=6)

        pr = ttk.Labelframe(right, text="Preview Waveform (.oto seleccionado)")
        pr.pack(fill="both", expand=True)

        self.pfig = Figure(figsize=(6.5, 4.3), dpi=100)
        self.pax = self.pfig.add_subplot(111)
        self.preview_canvas = FigureCanvasTkAgg(self.pfig, master=pr)
        self.preview_canvas.get_tk_widget().pack(fill="both", expand=True)

        self.prev_status = tk.StringVar(value="Selecciona un .oto para previsualizar.")
        ttk.Label(pr, textvariable=self.prev_status).pack(anchor="w", padx=8, pady=(0, 8))

    def pick_root(self):
        d = filedialog.askdirectory(title="Seleccionar carpeta raíz de waveforms")
        if d:
            self.root_dir.set(d)

    def clear_filters(self):
        self.fab_var.set("")
        self.marca_var.set("")
        self.modelo_var.set("")
        self.year_var.set("")
        self.engine_var.set("")
        self.module_var.set("")
        self.search_text.set("")
        self.refresh_filters()
        self.refresh_list()

    def on_filter_change(self):
        self.refresh_filters()
        self.refresh_list()

    def refresh_filters(self):
        where = {
            "fabricante": self.fab_var.get().strip(),
            "marca": self.marca_var.get().strip(),
            "modelo": self.modelo_var.get().strip(),
            "year": self.year_var.get().strip(),
            "engine": self.engine_var.get().strip(),
            "module": self.module_var.get().strip(),
        }

        fabs = [""] + self.db.list_distinct("fabricante", {})
        self.cb_fab["values"] = fabs
        if self.fab_var.get() not in fabs:
            self.fab_var.set("")

        marcas = [""] + self.db.list_distinct("marca", {"fabricante": where["fabricante"]} if where["fabricante"] else {})
        self.cb_marca["values"] = marcas
        if self.marca_var.get() not in marcas:
            self.marca_var.set("")

        dep = {}
        if self.fab_var.get():
            dep["fabricante"] = self.fab_var.get()
        if self.marca_var.get():
            dep["marca"] = self.marca_var.get()
        modelos = [""] + self.db.list_distinct("modelo", dep)
        self.cb_modelo["values"] = modelos
        if self.modelo_var.get() not in modelos:
            self.modelo_var.set("")

        dep2 = dict(dep)
        if self.modelo_var.get():
            dep2["modelo"] = self.modelo_var.get()
        years = [""] + self.db.list_distinct("year", dep2)
        self.cb_year["values"] = years
        if self.year_var.get() not in years:
            self.year_var.set("")

        dep3 = dict(dep2)
        if self.year_var.get():
            dep3["year"] = self.year_var.get()
        engines = [""] + self.db.list_distinct("engine", dep3)
        self.cb_engine["values"] = engines
        if self.engine_var.get() not in engines:
            self.engine_var.set("")

        dep4 = dict(dep3)
        if self.engine_var.get():
            dep4["engine"] = self.engine_var.get()
        modules = [""] + self.db.list_distinct("module", dep4)
        self.cb_module["values"] = modules
        if self.module_var.get() not in modules:
            self.module_var.set("")

    def refresh_list(self):
        fab = self.fab_var.get().strip()
        marca = self.marca_var.get().strip()
        modelo = self.modelo_var.get().strip()
        year = self.year_var.get().strip()
        engine = self.engine_var.get().strip()
        module = self.module_var.get().strip()
        text = self.search_text.get().strip()

        self.records = self.db.query(
            fabricante=fab, marca=marca, modelo=modelo, year=year, engine=engine, module=module, text=text
        )

        for item in self.tree.get_children():
            self.tree.delete(item)

        for rec in self.records:
            self.tree.insert("", "end", iid=str(rec.id),
                             values=(rec.profile, rec.fabricante, rec.year, rec.marca, rec.modelo, rec.engine, rec.module, rec.file_name))

        self.prev_status.set(f"Resultados: {len(self.records)}")

        self.selected_rec = None
        self._clear_preview()

    def _clear_preview(self):
        self.pax.clear()
        self.pax.set_title("Preview")
        self.preview_canvas.draw()

    def scan_root(self):
        root = self.root_dir.get().strip()
        if not root or not os.path.isdir(root):
            messagebox.showerror("Escaneo", "Carpeta raíz inválida.")
            return

        n = 0
        t0 = time.time()
        try:
            for dirpath, _, filenames in os.walk(root):
                for fn in filenames:
                    if not fn.lower().endswith(".oto"):
                        continue
                    fp = os.path.join(dirpath, fn)

                    file_meta = parse_oto_metadata_fast(fp)
                    path_meta = infer_meta_from_path(root, fp)
                    meta = merge_meta(file_meta, path_meta)

                    self.db.upsert(
                        profile=meta.get("profile", os.path.splitext(fn)[0]),
                        fabricante=meta.get("fabricante", ""),
                        marca=meta.get("marca", ""),
                        modelo=meta.get("modelo", ""),
                        year=meta.get("year", ""),
                        engine=meta.get("engine", ""),
                        module=meta.get("module", ""),
                        file_path=fp
                    )
                    n += 1
        except Exception as e:
            messagebox.showerror("Escaneo", f"Error escaneando.\n\n{e}")
            return

        dt = time.time() - t0
        messagebox.showinfo("Escaneo", f"Listo.\nArchivos procesados: {n}\nTiempo: {dt:.1f}s")
        self.refresh_filters()
        self.refresh_list()

    def on_select_row(self, event=None):
        sel = self.tree.selection()
        if not sel:
            self.selected_rec = None
            self._clear_preview()
            return

        try:
            rid = int(sel[0])
        except Exception:
            return
        rec = next((r for r in self.records if r.id == rid), None)
        self.selected_rec = rec
        if not rec:
            self._clear_preview()
            return

        self.prev_status.set(f"Seleccionado: {rec.file_name}")
        self.draw_preview(rec.file_path)

    def draw_preview(self, path: str):
        c = self.theme.colors()

        self.pax.clear()
        self.pfig.patch.set_facecolor(c["panel"])
        self.pax.set_facecolor(c["panel"])

        if not path or not os.path.isfile(path):
            self.pax.set_title("Preview: (archivo no encontrado)")
            self.preview_canvas.draw()
            return

        try:
            data = read_waveform_from_oto(path)
        except Exception as e:
            self.pax.set_title("Preview: error leyendo")
            self.pax.text(0.5, 0.5, str(e), ha="center", va="center", transform=self.pax.transAxes)
            self.preview_canvas.draw()
            return

        degs = list(range(TOTAL_DEGREES))
        offset = 0.0
        offset_step = 1.2
        plotted = 0

        enabled = self.editor_tab.model.channel_enabled

        for ch in CHANNELS:
            if not enabled.get(ch, True):
                continue
            vals = [v + offset for v in data.get(ch, [0] * TOTAL_DEGREES)]
            self.pax.plot(degs, vals, label=ch)
            plotted += 1
            offset += offset_step

        self.pax.set_title(os.path.basename(path))
        self.pax.set_xlabel("Degree")
        self.pax.set_ylabel("Nivel (desplazado)")
        self.pax.set_xlim(0, TOTAL_DEGREES - 1)

        xt = list(range(0, TOTAL_DEGREES, 50))
        if (TOTAL_DEGREES - 1) not in xt:
            xt.append(TOTAL_DEGREES - 1)
        self.pax.set_xticks(xt)

        if plotted > 0:
            max_offset = (plotted - 1) * offset_step
            self.pax.set_ylim(-0.5, max_offset + 1.8)
            self.pax.legend(loc="upper right", fontsize="x-small")
        else:
            self.pax.set_ylim(-0.5, 1.8)

        try:
            self.pfig.tight_layout()
        except Exception:
            pass

        self.preview_canvas.draw()

    def load_selected_to_editor(self):
        if not self.selected_rec:
            messagebox.showwarning("Editor", "Selecciona un registro primero.")
            return
        fp = self.selected_rec.file_path
        if not os.path.isfile(fp):
            messagebox.showerror("Editor", "El archivo no existe en disco.")
            return
        try:
            self.editor_tab.model.push_undo_snapshot()
            self.editor_tab.model.data = read_waveform_from_oto(fp)
            self.editor_tab.model.last_oto_path = fp
            self.app.notebook.select(self.app.plot_tab)
            self.editor_tab.redraw_plot()
        except Exception as e:
            messagebox.showerror("Editor", f"No se pudo cargar.\n\n{e}")

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Eliminar", "Selecciona uno o más registros.")
            return
        if not messagebox.askyesno("Eliminar", f"¿Eliminar {len(sel)} registros de la DB?\n(No borra archivos .oto del disco)"):
            return
        ids = []
        for s in sel:
            try:
                ids.append(int(s))
            except Exception:
                pass
        self.db.delete_ids(ids)
        self.refresh_filters()
        self.refresh_list()

    def edit_selected(self):
        if not self.selected_rec:
            messagebox.showwarning("Editar", "Selecciona un registro primero.")
            return

        rec = self.selected_rec

        win = tk.Toplevel(self)
        win.title("Editar metadata")
        win.transient(self.app)
        win.grab_set()

        frm = ttk.Frame(win)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        def row(lbl, val):
            r = ttk.Frame(frm)
            r.pack(fill="x", pady=4)
            ttk.Label(r, text=lbl, width=14).pack(side="left")
            v = tk.StringVar(value=val)
            e = ttk.Entry(r, textvariable=v)
            e.pack(side="left", fill="x", expand=True)
            return v

        v_prof = row("Profile", rec.profile)
        v_fab = row("Manuf", rec.fabricante)
        v_mar = row("Make", rec.marca)
        v_mod = row("Model", rec.modelo)
        v_year = row("Year", rec.year)
        v_eng = row("Engine", rec.engine)
        v_module = row("Module", rec.module)

        btn = ttk.Frame(frm)
        btn.pack(fill="x", pady=(10, 0))
        ttk.Button(btn, text="Cancelar", command=win.destroy).pack(side="right")
        ttk.Button(
            btn, text="Guardar",
            command=lambda: self._save_edit(win, rec.id, v_prof, v_fab, v_mar, v_mod, v_year, v_eng, v_module)
        ).pack(side="right", padx=6)

    def _save_edit(self, win, rid: int, v_prof, v_fab, v_mar, v_mod, v_year, v_eng, v_module):
        self.db.update_record(
            rec_id=rid,
            profile=v_prof.get().strip(),
            fabricante=v_fab.get().strip(),
            marca=v_mar.get().strip(),
            modelo=v_mod.get().strip(),
            year=v_year.get().strip(),
            engine=v_eng.get().strip(),
            module=v_module.get().strip()
        )
        win.destroy()
        self.refresh_filters()
        self.refresh_list()


# ---------------------------- Tab Configuración ----------------------------

class SettingsTab(ttk.Frame):
    def __init__(self, master, app, theme: ThemeManager):
        super().__init__(master)
        self.app = app
        self.theme = theme
        self._build_ui()

    def _build_ui(self):
        outer = ttk.Frame(self)
        outer.pack(fill="both", expand=True, padx=14, pady=14)

        lf = ttk.Labelframe(outer, text="Tema")
        lf.pack(fill="x")

        row = ttk.Frame(lf)
        row.pack(fill="x", padx=10, pady=10)

        ttk.Label(row, text="Selecciona tema:").pack(side="left")
        cb = ttk.Combobox(row, textvariable=self.theme.theme_var, state="readonly",
                          values=["Oscuro", "Claro"], width=12)
        cb.pack(side="left", padx=8)
        ttk.Button(row, text="Aplicar", command=self.apply_theme).pack(side="left")

        hint = ttk.Labelframe(outer, text="Notas")
        hint.pack(fill="x", pady=(12, 0))
        ttk.Label(
            hint,
            text=(
                "Si algo se ve raro al cambiar de tema, reinicia la app.\n"
                "Para hacer .exe: instala pyinstaller y ejecuta:\n"
                "    pip install pyinstaller\n"
                "    pyinstaller --onefile --windowed waveform.py"
            )
        ).pack(anchor="w", padx=10, pady=10)

    def apply_theme(self):
        self.theme.apply()
        self.app.plot_tab.redraw_plot()
        if self.app.library_tab.selected_rec:
            self.app.library_tab.draw_preview(self.app.library_tab.selected_rec.file_path)


# ---------------------------- Main App ----------------------------

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Waveform workbench V18.01")
        self.minsize(1200, 720)

        self.theme = ThemeManager(self)
        self.theme.apply()

        app_dir = os.path.join(os.path.expanduser("~"), ".waveform_app")
        db_path = os.path.join(app_dir, "waveforms.db")
        self.db = WaveDB(db_path)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)

        self.plot_tab = PlotAndToolsTab(self.notebook, self, self.theme)
        self.image_tab = ImageTab(self.notebook, self, self.theme)
        self.library_tab = LibraryTab(self.notebook, self, self.theme, self.db, self.plot_tab)
        self.settings_tab = SettingsTab(self.notebook, self, self.theme)

        self.notebook.add(self.plot_tab, text="Gráfica + Tools")
        self.notebook.add(self.image_tab, text="Imagen")
        self.notebook.add(self.library_tab, text="Librería")
        self.notebook.add(self.settings_tab, text="Configuración")

        self._build_menu()
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_menu(self):
        m = tk.Menu(self)
        self.config(menu=m)

        filem = tk.Menu(m, tearoff=0)
        m.add_cascade(label="Archivo", menu=filem)
        filem.add_command(label="Importar .oto (Editor)", command=self.plot_tab.on_import_oto)
        filem.add_command(label="Escribir waveform en .oto", command=self.plot_tab.on_write_oto)
        filem.add_separator()
        filem.add_command(label="Salir", command=self.on_close)

        libm = tk.Menu(m, tearoff=0)
        m.add_cascade(label="Librería", menu=libm)
        libm.add_command(label="Escanear / Actualizar DB", command=self.library_tab.scan_root)
        libm.add_command(label="Cargar seleccionado al Editor", command=self.library_tab.load_selected_to_editor)

        viewm = tk.Menu(m, tearoff=0)
        m.add_cascade(label="Vista", menu=viewm)
        viewm.add_command(label="Tema Oscuro", command=lambda: self._set_theme("Oscuro"))
        viewm.add_command(label="Tema Claro", command=lambda: self._set_theme("Claro"))

        helpm = tk.Menu(m, tearoff=0)
        m.add_cascade(label="Ayuda", menu=helpm)
        helpm.add_command(label="Dependencias", command=self._show_deps)

    def _set_theme(self, name: str):
        self.theme.theme_var.set(name)
        self.theme.apply()
        self.plot_tab.redraw_plot()
        if self.library_tab.selected_rec:
            self.library_tab.draw_preview(self.library_tab.selected_rec.file_path)

    def _show_deps(self):
        msg = (
            "Dependencias:\n"
            "  - matplotlib\n"
            "  - pillow\n"
            "  - openpyxl (solo si exportas Excel)\n\n"
            "Instalar:\n"
            "  pip install matplotlib pillow openpyxl\n"
        )
        messagebox.showinfo("Dependencias", msg)

    def on_close(self):
        try:
            self.db.close()
        except Exception:
            pass
        self.destroy()


def run():
    try:
        app = MainApp()
        app.mainloop()
    except Exception:
        traceback.print_exc()
        try:
            messagebox.showerror("Error", traceback.format_exc())
        except Exception:
            pass


if __name__ == "__main__":
    run()
